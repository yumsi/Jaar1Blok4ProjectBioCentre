########################################################################
# Author:Han Teunissen and Harm Laurense
# last update: 13-6-2019
# Function: Inserts a part of the data from the XML files into the database
# Known bugs: The script is heavily dependent on the NCBI servers so run
# this on a time where the servers aren't too busy
########################################################################
from Bio.Blast import NCBIXML
from tkinter import filedialog
from openpyxl import load_workbook
import re
import mysql.connector
from Bio import SearchIO, Entrez, SeqIO
import urllib
import time

def main():
    fname = filedialog.askopenfilename(title='Select')
    hdr, seqs = file_reader(fname)
    lijst_posities, data = xml_file_reader(hdr, seqs)
    data = data_sorter(hdr, lijst_posities, data)
    data_tuplelist_protein = data_sorteren(data)
    organism(data_tuplelist_protein)


def file_reader(fname):
    """ The file_reader reads the Excel file and gets the header and the
    sequence and puts them both in a separate lists(see return).
    :param fname: name/location of the Eccel file
    :return: hdr: list with the headers of the Excel file
    seqs: list with the sequences of the Excel file
    """
    hdr = []
    seqs = []
    # Loads dataset from Excel
    dataset_alles = load_workbook(fname)
    # Gets the data from groep 8
    dataset_groep8 = dataset_alles['groep8']
    for column_seq in dataset_groep8.iter_cols(min_row=1, max_col=5,
                                               max_row=100,
                                               values_only=True):
        for seq in column_seq:
            if seq.startswith('@HWI-'):
                seq = seq.replace('-', '').replace('@', '').replace(':', '')
                hdr.append(seq)
            elif not re.search('[^ATGCN]', seq):
                seqs.append(seq)
    return hdr, seqs


def xml_file_reader(hdr, seqs):
    """The xml_file_reader function reads the XML by looping through the hdr
    list. Then the function puts all the data in a list(see data). The
    function also makes a list with the position  of the data
    :param hdr: list with the headers of the Excel file
    :param seqs: list with the sequences of the Excel file
    :return: lijst_posities: list with positions of the data
    data: a list with the data form the XML files
    *note: The XML have to be in the same directory as the py file
    """
    data = []
    lijst_posities = []
    positie = -1
    positie2 = -1
    for header in hdr:
        positie2 += 1
        result_handle = open(header)
        blast_record = NCBIXML.read(result_handle)
        count = 0
        for alignment in blast_record.alignments:
            for hsp in alignment.hsps:
                if count < 100:
                    positie += 1
                    # appends the position to lijst_posities
                    lijst_posities.append(positie)
                    data.append([])
                    count += 1
                    query_coverage = round((hsp.query_end - hsp.query_start)
                                           / hsp.query_end, 3)
                    # appends information we want to a list
                    data[positie].append(hdr[positie2])
                    data[positie].append(seqs[positie2])
                    data[positie].append(hsp.expect)
                    data[positie].append(hsp.score)
                    data[positie].append(hsp.identities)
                    data[positie].append(query_coverage)
                    data[positie].append(hsp.sbjct)
    return lijst_posities, data


def data_sorter(hdr, lijst_posities, data):
    """ The data_sorter function sorts the data from the XML files. It also
    ''cleans'' the data so it looks better in the database.
    :param hdr: list with the headers of the Excel file
    :param lijst_posities:  list with positions of the data
    :param data: a list with the data form the XML files
    :return: data: tuples of filtered data from the XML files
    """
    teller = -1
    hdr_filtered = []
    for header in hdr:
        query_result = SearchIO.parse(header, 'blast-xml')
        for result in query_result:
            count = 0
            for hit in result:
                # checks every chosen hit
                for hsp in hit:
                    if count < 100:
                        teller += 1
                        data[lijst_posities[teller]].append(
                            hit.description.split('[')[1].strip(']'))
                        data[lijst_posities[teller]].append(
                            hit.description.split('[')[0])
                        data[lijst_posities[teller]].append(hit.accession)
                        count += 1
    for header in hdr:
        blast_qresult = SearchIO.read(header, "blast-xml")
        if len(blast_qresult) != 0:
            hdr_filtered.append(header)
    return data


def data_sorteren(data):
    """ The data_sorteren sorts the data and makes them into tuples to
    insert into the database.
    :param data: tuples of filtered data from the XML files
    :return: data_tuplelist_lineage: Tuplelist of the lineage to insert into
    the database
     data_tuplelist_protein: Tuplelist of the protein to insert into
    the database
     data_tuplelist_fragment: Tuplelist of the fragment to insert into
    the database
    """
    data_list_protein = []
    data_tuplelist_protein = []
    counter = -1
    for lijst in data:
        data_list_protein.append((lijst[2:7]))
    for lijst2 in data:
        counter += 1
        data_list_protein[counter].extend(tuple(lijst2[8:10]))
    for lijst3 in data_list_protein:
        data_tuplelist_protein.append(tuple(lijst3))
    return data_tuplelist_protein


def organism(data_tuplelist_protein):
    """The organism function gets the taxonomy, taxonomy id and protein
    description from the NCBI website by using the biopython module. This
    function also detects if the is an error in the connection with the NCBI
    server and tries to reestablish the connection to the NCBI server.
    :param data_tuplelist_protein: A list with the accesioncodes of the
    proteins
    :return: A print statement when the code has finished succesfully
    """
    lijst_gemistehits = []
    lijst_gemistehits_posities = []
    Entrez.email = "probaly_school_stuff@outlook.com"
    print("gaat even duren....")
    counter = -1
    for code in data_tuplelist_protein:
        if counter < 10000:
            try:
                time.sleep(0.3)
                data_list_taxonomie = []
                data_list_eiwitdescriptie = []
                counter += 1
                handle = Entrez.efetch(db="protein", id=code[6], rettype="gb",
                                       retmode="text")
                record = SeqIO.read(handle, "genbank")
                taxo_string = ""
                taxo = record.annotations['taxonomy']
                for element in taxo:
                    taxo_string = taxo_string + element + "; "
                data_list_taxonomie.append([taxo_string])
                if record.annotations['taxonomy'] == []:
                    tax_id = [""]
                    data_list_taxonomie[0].extend(tax_id)
                    desc = ""
                    data_list_eiwitdescriptie.append(desc)
                    data_insertie2(data_list_taxonomie,
                                   data_list_eiwitdescriptie)
                else:
                    species = record.annotations['taxonomy'][-1]
                    tax_id = taxo_id(species)
                    data_list_taxonomie[0].append(tax_id[0])
                    desc = description(record)
                    desc = desc
                    data_list_eiwitdescriptie.append(desc)
                    data_insertie2(data_list_taxonomie,
                                   data_list_eiwitdescriptie)
            except (urllib.error.HTTPError, urllib.error.URLError,
                    ConnectionResetError):
                try:
                    time.sleep(10)
                    data_list_taxonomie = []
                    data_list_eiwitdescriptie = []
                    print(counter)
                    handle = Entrez.efetch(db="protein", id=code[6],
                                           rettype="gb", retmode="text")
                    record = SeqIO.read(handle, "genbank")
                    taxo_string = ""
                    taxo = record.annotations['taxonomy']
                    for element in taxo:
                        taxo_string = taxo_string + element + "; "
                    data_list_taxonomie.append([taxo_string])
                    if record.annotations['taxonomy'] == []:
                        tax_id = [""]
                        data_list_taxonomie[0].extend(tax_id)
                        desc = ""
                        data_list_eiwitdescriptie.append(desc)
                        data_insertie2(data_list_taxonomie,
                                       data_list_eiwitdescriptie)
                    else:
                        species = record.annotations['taxonomy'][-1]
                        tax_id = taxo_id(species)
                        data_list_taxonomie[0].append(tax_id[0])
                        desc = description(record)
                        desc = desc
                        data_list_eiwitdescriptie.append(desc)
                        data_insertie2(data_list_taxonomie,
                                       data_list_eiwitdescriptie)
                    lijst_gemistehits.append(code)
                    lijst_gemistehits_posities.append(counter)
                    continue
                except (urllib.error.HTTPError, urllib.error.URLError,
                        ConnectionResetError):
                    continue
    return print("Klaar met vullen database")


def taxo_id(species):
    """The taxo_id function gets the taxonomy id from the NCBI wensite using
    the biopython module. And sends this information back to the organism
    function.
    :param species: A list of the names of the organisms
    :return: tax_id: A list of the taxonomy id's from the NCBI website
    """
    try:
        handle = Entrez.esearch(db="taxonomy", term=species)
        record = Entrez.read(handle)
        tax_id = record['IdList']
        return tax_id
    except RuntimeError:
        try:
            time.sleep(10)
            handle = Entrez.esearch(db="taxonomy", term=species)
            record = Entrez.read(handle)
            tax_id = record['IdList']
            return tax_id
        except RuntimeError:
            tax_id = [""]
            return tax_id


def description(record):
    """The description function gets the information from the protein file
    in the NCBI protein database. And sends a list of the description back
    to the organism function.
    :param record: A list of the protein records from the NCBI databse
    :return: desc: A list of the description of the proteins
    """
    try:
        for line in record.annotations['references']:
            line = str(line)
            # splits the line so only the desription remains
            extra1, title1 = line.split('title:')
            desc, extra = title1.split('journal')
            desc = desc.strip()
            return desc
    except KeyError:
        desc = ''
        return desc


def data_insertie2(data_list_taxonomie, data_list_eiwitdescriptie):
    """The data_insertie2 functon puts the information from the tuples(see
    parameters) into the database.

    :param data_list_taxonomie: Tuples with information of the taxonomy of
    the organisms of the proteins
    :param data_list_eiwitdescriptie: Tuples with information of the
    taxonomy of the description of the protein
    :return: a print when everything is done.
    """
    data_tuplelist_taxonomie = []
    data_tuplelist_eiwitdescriptie = []
    for list in data_list_taxonomie:
        data_tuplelist_taxonomie.append(tuple(list))
    for list2 in data_list_eiwitdescriptie:
        data_tuplelist_eiwitdescriptie.append(tuple([list2]))
    # connection information
    conn = mysql.connector.connect(
        host="hannl-hlo-bioinformatica-mysqlsrv.mysql.database.azure.com",
        user="yumsi@hannl-hlo-bioinformatica-mysqlsrv",
        password="yumsi11",
        db="yumsi")
    # query for the database
    query_taxonomie = """INSERT INTO ProjectBlok4_Taxonomy (Taxonomie,  \
                      Accessiecode_taxonomie) VALUES (%s, %s) """
    insert_gegevens_taxonomie = data_tuplelist_taxonomie
    cursor = conn.cursor(prepared=True)
    cursor.executemany(query_taxonomie, insert_gegevens_taxonomie)
    cursor.close()
    conn.commit()
    print("---------Data insertion into table Lineage done----------")
    query_functie = "INSERT INTO ProjectBlok4_Function (Functie_naam) " \
            "VALUES (%s)"
    insert_gegevens_functie = data_tuplelist_eiwitdescriptie
    cursor = conn.cursor(prepared=True)
    cursor.executemany(query_functie, insert_gegevens_functie)
    cursor.close()
    conn.commit()
    print("---------Data insertion into table Function done----------")
    print("Insertion done, closing connection...")
    conn.close()
    return print("Connection closed")


main()