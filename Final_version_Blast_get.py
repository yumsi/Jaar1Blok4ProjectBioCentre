########################################################################
# Author:Han Teunissen and Harm Laurense
# last update: 13-6-2019
# Function: Gets the  BLAST files and makes them into an XML file
# Known bugs: The application gets blastp results even though the files says
# the results are from blastp
########################################################################
from Bio.Blast import NCBIWWW
from tkinter import filedialog
from openpyxl import load_workbook
import re
import time
from tkinter import messagebox


def main():
    # File selection for the user
    fname = filedialog.askopenfilename(
        initialdir='C:\\Users\\hante\\.PyCharm2019.1\\config\\scratches\\Test '
                   'files',
        title='Select')
    hdr, seqs = file_reader(fname)
    blaster_file(hdr, seqs, count=0)


def file_reader(fname):
    """ The file_reader reads the Excel file and gets the header and the
    sequence and puts them both in a separate lists(see return).
    :param fname: name/location of the Eccel file
    :return: hdr: list with the headers of the Excel file
    seqs: list with the sequences of the Excel file
    """
    hdr = []
    seqs = []
    # Loads dataset from Excel
    dataset_alles = load_workbook(fname)
    # Gets the data from groep 8
    dataset_groep8 = dataset_alles['groep8']
    for column_seq in dataset_groep8.iter_cols(min_row=1, max_col=5,
                                               max_row=100,
                                               values_only=True):
        for seq in column_seq:
            if seq.startswith('@HWI-'):
                seq = seq.replace('-', '').replace('@', '').replace(':', '')
                hdr.append(seq)
            elif not re.search('[^ATGCN]', seq):
                seqs.append(seq)
    return hdr, seqs


def blaster_file(hdr, seqs, count):
    """ The blaster_file function gets every sequence and then uses the
    online BLASTx tool (from the NCBI) to blast every sequence. The results
    from these blasts are then put into a file with the appropriate sequence
    header. Furthermore it notes a number in a separate file to keep track
    at wich seqence it was in case of an error.
    :param hdr: list with the headers of the Excel file
    :param seqs: list with the sequences of the Excel file
    :param count: the last index number in the seqs list
    :return: A tkinter message *if everything did go correctly) when all the
    sequences are BLASTED
    """
    try:
        for seq in seqs:
            count = int(count)
            # using time.sleep to not get kicked out of the NCBI server
            time.sleep(10)
            # Blasting the sequence online against blastx
            result_handle = NCBIWWW.qblast("blastx", "nr", seq)
            count = int(count)
            header = hdr[count]
            # Makes the XML with the appropriate header
            file = open(header, "x")
            file.write(result_handle.read())
            file.close()
            # creatig a back-up number
            back_up = open('Blast_text', 'w')
            back_up.write(str(count))
            back_up.close()
            count += 1
    except FileExistsError:
        # error so you do not redownload the same file
        file = open('Blast_text')
        count = file.readline()
        file.close()
        # checks for the back-up number
        file = open('Blast_text', 'w')
        count = int(count)
        count += 1
        file.write(str(count))
        file.close()
        # returns to the function with the new back-p number
        return blaster_file(hdr, seqs, count)
    except IndexError:
        messagebox.showinfo('Error', 'Alle bestanden zijn aangemaakt.')


main()
