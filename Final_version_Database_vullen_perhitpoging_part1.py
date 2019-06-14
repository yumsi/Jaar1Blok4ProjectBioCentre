########################################################################
# Author:Han Teunissen and Harm Laurense
# last update: 13-6-2019
# Function: Inserts a part of the data from the XML files into the database
# Known bugs: The script is heavily dependent on the NCBI servers so run
# this on a time where the servers aren't too busy
########################################################################
from Bio.Blast import NCBIXML
from tkinter import filedialog
from openpyxl import load_workbook
import re
import mysql.connector
from Bio import SearchIO


def main():
    fname = filedialog.askopenfilename(title='Select')
    hdr, seqs = file_reader(fname)
    lijst_posities, data = xml_file_reader(hdr, seqs)
    data = data_sorter(hdr, lijst_posities, data)
    data_tuplelist_lineage, data_tuplelist_protein, data_tuplelist_fragment \
        = data_sorteren(data)
    data_insertie(data_tuplelist_lineage, data_tuplelist_protein,
                  data_tuplelist_fragment)


def file_reader(fname):
    """ The file_reader reads the Excel file and gets the header and the
    sequence and puts them both in a separate lists(see return).
    :param fname: name/location of the Eccel file
    :return: hdr: list with the headers of the Excel file
    seqs: list with the sequences of the Excel file
    """
    hdr = []
    seqs = []
    # Loads dataset from Excel
    dataset_alles = load_workbook(fname)
    # Gets the data from groep 8
    dataset_groep8 = dataset_alles['groep8']
    for column_seq in dataset_groep8.iter_cols(min_row=1, max_col=5,
                                               max_row=100,
                                               values_only=True):
        for seq in column_seq:
            if seq.startswith('@HWI-'):
                seq = seq.replace('-', '').replace('@', '').replace(':', '')
                hdr.append(seq)
            elif not re.search('[^ATGCN]', seq):
                seqs.append(seq)
    return hdr, seqs


def xml_file_reader(hdr, seqs):
    """The xml_file_reader function reads the XML by looping through the hdr
    list. Then the function puts all the data in a list(see data). The
    function also makes a list with the position  of the data
    :param hdr: list with the headers of the Excel file
    :param seqs: list with the sequences of the Excel file
    :return: lijst_posities: list with positions of the data
    data: a list with the data form the XML files
    *note: The XML have to be in the same directory as the py file
    """
    data = []
    lijst_posities = []
    positie = -1
    positie2 = -1
    for header in hdr:
        positie2 += 1
        result_handle = open(header)
        blast_record = NCBIXML.read(result_handle)
        count = 0
        for alignment in blast_record.alignments:
            for hsp in alignment.hsps:
                if count < 100:
                    positie += 1
                    # appends the position to lijst_posities
                    lijst_posities.append(positie)
                    data.append([])
                    count += 1
                    query_coverage = round((hsp.query_end - hsp.query_start)
                                           / hsp.query_end, 3)
                    # appends information we want to a list
                    data[positie].append(hdr[positie2])
                    data[positie].append(seqs[positie2])
                    data[positie].append(hsp.expect)
                    data[positie].append(hsp.score)
                    data[positie].append(hsp.identities)
                    data[positie].append(query_coverage)
                    data[positie].append(hsp.sbjct)
    return lijst_posities, data


def data_sorter(hdr, lijst_posities, data):
    """ The data_sorter function sorts the data from the XML files. It also
    ''cleans'' the data so it looks better in the database.
    :param hdr: list with the headers of the Excel file
    :param lijst_posities:  list with positions of the data
    :param data: a list with the data form the XML files
    :return: data: tuples of filtered data from the XML files
    """
    teller = -1
    hdr_filtered = []
    for header in hdr:
        query_result = SearchIO.parse(header, 'blast-xml')
        for result in query_result:
            count = 0
            for hit in result:
                for hsp in hit:
                    if count < 100:
                        teller += 1
                        data[lijst_posities[teller]].append(
                            hit.description.split('[')[1].strip(']'))
                        data[lijst_posities[teller]].append(
                            hit.description.split('[')[0])
                        data[lijst_posities[teller]].append(hit.accession)
                        count += 1
    for header in hdr:
        blast_qresult = SearchIO.read(header, "blast-xml")
        if len(blast_qresult) != 0:
            hdr_filtered.append(header)
    return data


def data_sorteren(data):
    """ The data_sorteren sorts the data and makes them into tuples to
    insert into the database.
    :param data: tuples of filtered data from the XML files
    :return: data_tuplelist_lineage: Tuplelist of the lineage to insert into
    the database
     data_tuplelist_protein: Tuplelist of the protein to insert into
    the database
     data_tuplelist_fragment: Tuplelist of the fragment to insert into
    the database
    """
    data_tuplelist_lineage = []
    data_list_protein = []
    data_tuplelist_protein = []
    data_tuplelist_fragment = []
    counter = -1
    for lijst in data:
        data_tuplelist_fragment.append(tuple(lijst[0:2]))
        data_list_protein.append((lijst[2:7]))
        data_tuplelist_lineage.append(tuple([lijst[7]]))
    for lijst2 in data:
        counter += 1
        data_list_protein[counter].extend(tuple(lijst2[8:10]))
    for lijst3 in data_list_protein:
        data_tuplelist_protein.append(tuple(lijst3))
    return data_tuplelist_lineage, data_tuplelist_protein, \
           data_tuplelist_fragment


def data_insertie(data_tuplelist_lineage, data_tuplelist_protein,
                  data_tuplelist_fragment):
    """The data_insertie insert the tuple lists(se parameters) into the
    database using mysql.connector.
    :param data_tuplelist_lineage:  Tuplelist of the lineage to insert into
    the database
    :param data_tuplelist_protein: Tuplelist of the protein to insert into
    the database
    :param data_tuplelist_fragment:  Tuplelist of the fragment to insert into
    the database
    :return: message when the lists are succesfully inserted into the database
    """
    conn = mysql.connector.connect(
        host="hannl-hlo-bioinformatica-mysqlsrv.mysql.database.azure.com",
        user="yumsi@hannl-hlo-bioinformatica-mysqlsrv",
        password="yumsi11",
        db="yumsi")

    query = "INSERT INTO ProjectBlok4_Lineage (Lineage_naam) " \
            "VALUES (%s)"
    insert_gegevens = data_tuplelist_lineage
    # Prepared=True en executemany zorgen dat je meerdere lijsten aan gegevens
    # kan gebruiken
    cursor = conn.cursor(prepared=True)
    cursor.executemany(query, insert_gegevens)
    cursor.close()
    conn.commit()
    print("---------Data insertion into table Lineage done----------")

    query1 = "INSERT INTO ProjectBlok4_Fragment (Fragment_naam, " \
             "Fragment_sequentie) " \
             "VALUES (%s, %s)"
    insert_gegevens1 = data_tuplelist_fragment
    cursor = conn.cursor(prepared=True)
    cursor.executemany(query1, insert_gegevens1)
    cursor.close()
    conn.commit()
    print("---------Data insertion into table Fragment done----------")
    query2 = "INSERT INTO ProjectBlok4_Protein (Expect, " \
             "Alignment_scores, Per_ident, Query_coverage, Sequentie, " \
             "Eiwit_Naam, Accessiecode) " \
             "VALUES (%s, %s, %s, %s, %s, %s, %s)"
    insert_gegevens2 = data_tuplelist_protein
    cursor = conn.cursor(prepared=True)
    cursor.executemany(query2, insert_gegevens2)
    cursor.close()
    conn.commit()
    print("---------Data insertion into table Protein done----------")
    print("Insertion done, closing connection...")
    conn.close()
    print("Connection closed")


main()
